"""
광고 분석 API 라우트
- 17개 엔드포인트 구현
- 파일 업로드 및 데이터 처리
"""

import os
import pandas as pd
import numpy as np
import logging
from flask import (
    Blueprint, render_template, request, jsonify,
    session, redirect, url_for, send_file, send_from_directory, current_app, g
)
from werkzeug.utils import secure_filename
import flask

from app.services.ad_analyzer import AdAnalyzer
from app.services.ai_insights import AIInsights
from app.utils.db_utils import execute_query, execute_insert, execute_update, DatabaseError
from app.utils.helpers import (
    allowed_file, clean_filename, get_unique_filename,
    create_error_response, create_success_response,
    ensure_directory_exists
)


from itsdangerous import URLSafeTimedSerializer

logger = logging.getLogger(__name__)

# ========================================
# 소셜 미디어 봇 감지
# ========================================
BOT_USER_AGENTS = [
    # 글로벌 소셜/메신저
    'facebookexternalhit',  # Facebook
    'Facebot',              # Facebook
    'Twitterbot',           # Twitter/X
    'LinkedInBot',          # LinkedIn
    'Slackbot',             # Slack
    'TelegramBot',          # Telegram
    'WhatsApp',             # WhatsApp
    'Discordbot',           # Discord
    'Pinterest',            # Pinterest

    # 검색엔진
    'Googlebot',            # Google
    'bingbot',              # Bing

    # 네이버 관련
    'Yeti',                 # Naver 검색
    'naver.me',             # Naver 공통 식별자
    'NaverBot',             # Naver 봇
    'WorksOgCrawler',       # Naver Works OG 크롤러
    'naverbookmarkcrawler', # Naver 북마크
    'scrapbook-scraper',    # 스크랩북

    # 카카오/다음 관련
    'kakaotalk-scrap',      # KakaoTalk
    'Daumoa',               # Daum 검색
]

def is_social_bot(user_agent_string):
    """소셜 미디어 봇인지 확인 (OG 태그 크롤러)"""
    if not user_agent_string:
        return False
    ua_lower = user_agent_string.lower()
    return any(bot.lower() in ua_lower for bot in BOT_USER_AGENTS)

# ========================================
# 제외 키워드 판정 상수
# ========================================
EXCLUDE_MIN_SPEND = 5000      # 최소 광고비 (원)
EXCLUDE_MIN_CLICKS = 10       # 최소 클릭수
EXCLUDE_CPC_CRITICAL = 500    # CPC 심각 기준 (원)
EXCLUDE_CPC_VERY_HIGH = 800   # CPC 매우 높음 기준 (원)
EXCLUDE_CLICKS_CRITICAL = 30  # 전환없음 즉시제외 클릭수
EXCLUDE_CLICKS_HIGH = 15      # 전환없음 조속히제외 클릭수

# ========================================
# 임시 인증 함수 (TODO: 추후 재설계)
# ========================================
def get_current_user():
    user = {
        'userId': session.get('userId', 'test'),
        'userNicknm': session.get('userNicknm', 'testNicknm')
    }
    return user
    
def get_current_user_id():
    return session.get('userId', 'test')

# Blueprint 생성
ad_bp = Blueprint('ad_analysis', __name__)

# ========================================
# 컬럼 매핑 정의 (한글 → 영문)
# ========================================
COLUMN_MAPPING = {
    '날짜': 'date',
    '캠페인명': 'campaign_name',
    '광고유형': 'ad_type',
    '지출액': 'spend',
    '노출수': 'impressions',
    '클릭수': 'clicks',
    '전환수': 'conversions',
    '매출액': 'revenue'
}

# 광고유형 값 매핑 (한글 → 영문)
AD_TYPE_MAPPING = {
    '매출형': 'sales',
    '잠재고객': 'lead'
}


def normalize_columns(df):
    """
    한글/영문 컬럼을 자동 감지하여 영문으로 통일

    Args:
        df: pandas DataFrame

    Returns:
        DataFrame: 영문 컬럼명으로 통일된 DataFrame
    """
    # 한글 컬럼명을 영문으로 변환
    rename_map = {}
    for kor, eng in COLUMN_MAPPING.items():
        if kor in df.columns:
            rename_map[kor] = eng

    if rename_map:
        df = df.rename(columns=rename_map)
        logger.info(f'Column mapping applied: {rename_map}')

    # 광고유형 값 변환 (매출형 → sales, 잠재고객 → lead)
    if 'ad_type' in df.columns:
        df['ad_type'] = df['ad_type'].map(lambda x: AD_TYPE_MAPPING.get(x, x) if pd.notna(x) else 'sales')
        logger.info(f'Ad type values mapped: {df["ad_type"].value_counts().to_dict()}')
    else:
        # ad_type 컬럼이 없으면 기본값 'sales' 설정
        df['ad_type'] = 'sales'
        logger.info('No ad_type column found - defaulting to "sales"')

    return df



@ad_bp.before_app_request
def before_request():
    """
    요청 전 인증 체크
    
    - 개발 모드에서는 세션 체크를 건너뜀
    - 정적 파일 요청도 세션 체크 제외
    """
    # 정적 파일 및 공개 페이지는 세션 체크 제외
    if request.path.startswith('/static/') : return None
    if request.path.startswith('/landing'): return None

    # 소셜 미디어 봇이면 홈 페이지 세션 체크 건너뛰기 (OG 메타태그용)
    user_agent = request.headers.get('User-Agent', '')
    if request.path == '/' and is_social_bot(user_agent):
        return None  # index()에서 og_only.html 반환

    # 개발 모드 체크 (DEBUG 모드이거나 FLASK_ENV가 development인 경우)
    is_debug_mode = current_app.config.get('DEBUG', False)
    flask_env = current_app.config.get('FLASK_ENV', os.getenv('FLASK_ENV', 'development'))
    is_development = flask_env == 'development' or is_debug_mode
    
    # 개발 모드이면 세션 체크 건너뛰기
    if is_development:
        logger.debug(f"[개발 모드] 세션 체크 건너뛰기: {request.path}")
        g.user = {'userId': 'test', 'userNicknm': 'testNicknm'}
        return None
    
    # 운영 모드에서는 세션 체크 수행
    COOKIE_VALUE = request.cookies.get('mbiz_session')
    SECRET_KEY = current_app.config.get('SECRET_KEY')
    SALT = 'cookie-session' # Flask 기본값

    serializer = URLSafeTimedSerializer(
        secret_key=SECRET_KEY,
        salt=SALT,
        serializer=flask.json.tag.TaggedJSONSerializer(),
        signer_kwargs={'key_derivation': 'hmac', 'digest_method': 'sha1'} 
    )

    try:
        data = serializer.loads(COOKIE_VALUE)
        if 'userId' in data :
            g.user = data
            print('g.user: ', g.user)
        else:
            g.user = {
                'userId': '',
                'name': '',
                'userNicknm': ''
            }
            if request.path == '/' : return None
            if request.path.startswith('/guide'): return None
            return redirect('https://mbizsquare.com/#/login')
    except Exception as e:
        print("❌ 실패! 정확한 에러 원인:", e)
        return redirect('https://mbizsquare.com/#/login')

# ========================================
# 1. 메인 페이지 및 인증
# ========================================

@ad_bp.route('/landing')
def landing():
    """
    랜딩페이지 (공개, 로그인 불필요)

    쿠팡 광고 대시보드 홍보 랜딩페이지
    카카오톡 오픈채팅방 유입 목적

    Returns:
        HTML: 랜딩페이지
    """
    return send_from_directory('static/landing', 'index.html')


@ad_bp.route('/')
def index():
    """
    홈 대시보드 (마케팅광장 광고분석 대시보드 - 마광)

    Returns:
        HTML: 홈 대시보드 템플릿
    """
    # 소셜 미디어 봇(OG 크롤러)이면 메타태그만 있는 페이지 반환
    user_agent = request.headers.get('User-Agent', '')
    if is_social_bot(user_agent):
        return render_template('og_only.html')

    # 메인 프로젝트 세션에서 사용자 정보 가져오기
    user = get_current_user()

    return render_template('home_dashboard.html', user=user)


@ad_bp.route('/ad-dashboard')
def dashboard():
    """
    광고 분석 대시보드 메인 페이지 (일반+메타)

    Returns:
        HTML: 대시보드 템플릿
    """
    # 메인 프로젝트 세션에서 사용자 정보 가져오기
    user = get_current_user()

    return render_template('ad_dashboard_v2.html', user=user)


@ad_bp.route('/ad-dashboard/coupang-test')
def coupang_manual_test():
    """
    쿠팡 광고 수동 테스트 페이지

    Returns:
        HTML: 수동 테스트 템플릿
    """
    return render_template('manual_test.html')


@ad_bp.route('/ad-dashboard/coupang')
@ad_bp.route('/ad-dashboard-coupang')  # 별칭 라우트 추가
def coupang_dashboard():
    """
    쿠팡 광고 전용 대시보드

    Returns:
        HTML: 쿠팡 대시보드 템플릿
    """
    # 메인 프로젝트 세션에서 사용자 정보 가져오기
    user = get_current_user()

    return render_template('ad_dashboard_coupang.html', user=user)


@ad_bp.route('/ad-dashboard/profit-simulator')
def profit_simulator():
    """
    수익 시뮬레이터 페이지

    Returns:
        HTML: 수익 시뮬레이터 템플릿
    """
    user = get_current_user()
    return render_template('profit_simulator.html', user=user)


@ad_bp.route('/ad-dashboard/ad-efficiency')
def ad_efficiency():
    """
    광고 효율 진단 페이지

    Returns:
        HTML: 광고 효율 진단 템플릿
    """
    user = get_current_user()
    return render_template('ad_efficiency.html', user=user)


@ad_bp.route('/ad-dashboard/keyword-combiner')
def keyword_combiner():
    """
    키워드 조합기 페이지

    Returns:
        HTML: 키워드 조합기 템플릿
    """
    user = get_current_user()
    return render_template('keyword_combiner.html', user=user)


@ad_bp.route('/guide')
def guide():
    """
    이용안내 페이지

    Returns:
        HTML: 이용안내 템플릿
    """
    return render_template('guide.html')


@ad_bp.route('/login')
def login():
    """
    로그인 페이지 (JWT 토큰 없이 접근 시)
    """
    return render_template('login.html')


@ad_bp.route('/logout')
def logout():
    """
    로그아웃 - 메인 프로젝트 로그인 페이지로 리다이렉트
    """
    session.clear()
    logger.info("User logged out")
    return redirect(current_app.config.get('MAIN_LOGIN_URL', 'https://mbizsquare.com/login'))


# ========================================
# 2. 데이터 업로드 API
# ========================================

@ad_bp.route('/api/ad-analysis/upload', methods=['POST'])
def upload_data():
    """Excel/CSV 파일 업로드 및 분석 (데이터베이스 저장)"""
    user_id = get_current_user_id()

    if 'file' not in request.files:
        return jsonify({'success': False, 'error': '파일이 없습니다'}), 400

    file = request.files['file']

    if file.filename == '':
        return jsonify({'success': False, 'error': '파일명이 비어있습니다'}), 400

    try:
        # 파일 읽기
        if file.filename.endswith('.csv'):
            df = pd.read_csv(file)
        else:
            # Excel 파일인 경우, 자동으로 적절한 시트 찾기
            xl_file = pd.ExcelFile(file)

            # 시트 우선순위: 일별데이터 > 광고데이터 > 첫 번째 시트
            if '일별데이터' in xl_file.sheet_names:
                df = pd.read_excel(xl_file, sheet_name='일별데이터')
            elif '광고데이터' in xl_file.sheet_names:
                df = pd.read_excel(xl_file, sheet_name='광고데이터')
            else:
                # 첫 번째 시트 읽기 (입력양식 시트 우선)
                if '입력양식' in xl_file.sheet_names:
                    df = pd.read_excel(xl_file, sheet_name='입력양식')
                else:
                    df = pd.read_excel(xl_file, sheet_name=0)

        # 컬럼 정규화 (한글 → 영문 변환 + 광고유형 처리)
        df = normalize_columns(df)

        # 필수 컬럼 확인 (ad_type은 normalize_columns에서 자동 추가됨)
        required_cols = ['date', 'campaign_name', 'spend', 'clicks', 'conversions', 'revenue']
        missing_cols = [col for col in required_cols if col not in df.columns]

        if missing_cols:
            # 한글 컬럼명으로 에러 메시지 표시
            kor_missing = [k for k, v in COLUMN_MAPPING.items() if v in missing_cols]
            return jsonify({'success': False, 'error': f'필수 컬럼 누락: {kor_missing or missing_cols}'}), 400

        # Impression 데이터 처리 (없거나 0이면 추정)
        impressions_estimated = False
        if 'impressions' not in df.columns or df['impressions'].sum() == 0:
            # CTR 2% 가정하여 노출수 추정
            df['impressions'] = (df['clicks'] * 50).astype(int)
            impressions_estimated = True
            logger.info('Impressions column missing or zero - estimated from clicks (CTR ~2%)')

        # 스냅샷 이름 생성
        snapshot_name = request.form.get('snapshot_name', f'업로드 {pd.Timestamp.now().strftime("%Y-%m-%d %H:%M")}')

        # 임시 스냅샷 ID 생성 (DB 대신 메모리 사용)
        snapshot_id = int(pd.Timestamp.now().timestamp())

        # In-Memory 방식으로 지표 계산 (DB 없이)
        metrics = _calculate_metrics_inmemory(df)

        # Add impression estimation flag to metrics
        metrics['impressions_estimated'] = impressions_estimated

        # AI 인사이트 생성 (선택사항)
        try:
            ai = AIInsights()
            insights = ai.generate_insights(metrics, df)
        except Exception as ai_error:
            logger.warning(f'AI insights generation failed: {ai_error}')
            insights = '✅ 분석 완료! 데이터가 성공적으로 처리되었습니다.'

        # 세션에 저장 (선택사항)
        session[f'snapshot_{snapshot_id}'] = {
            'name': snapshot_name,
            'data': df.to_dict('records'),
            'metrics': metrics,
            'insights': insights,
            'created_at': pd.Timestamp.now().isoformat()
        }

        logger.info(f'File uploaded and processed in-memory: {file.filename}, snapshot_id: {snapshot_id}')

        return jsonify({
            'success': True,
            'snapshot_id': snapshot_id,
            'metrics': metrics,
            'insights': insights
        })

    except Exception as e:
        logger.error(f'File upload failed: {e}')
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'error': f'처리 중 오류: {str(e)}'}), 500


@ad_bp.route('/api/ad-analysis/upload-coupang', methods=['POST'])
def upload_coupang():
    """
    쿠팡 광고 Excel 파일 업로드 및 파싱

    쿠팡 광고 보고서 필수 컬럼:
    - 키워드, 노출수, 클릭수, 광고비, 클릭률
    - 총 주문수(1일), 총 판매수량(1일), 총 전환매출액(1일)
    - 총광고수익률(1일) = ROAS
    """
    user_id = get_current_user_id()

    if 'file' not in request.files:
        return jsonify({'success': False, 'error': '파일이 없습니다'}), 400

    file = request.files['file']

    if file.filename == '':
        return jsonify({'success': False, 'error': '파일명이 비어있습니다'}), 400

    try:
        # Excel 파일 읽기 (인코딩 문제 해결 - BytesIO 사용)
        import io
        file_content = file.read()
        df = pd.read_excel(io.BytesIO(file_content), engine='openpyxl')
        logger.info(f'Coupang file uploaded: {file.filename}, rows: {len(df)}, columns: {len(df.columns)}')

        # 필수 컬럼 확인 (매출액은 14일 우선, 없으면 1일 사용)
        required_cols_base = ['키워드', '노출수', '클릭수', '광고비', '클릭률']

        # 경고 메시지 초기화
        warning_message = None
        data_type = '14일'

        # 매출액 컬럼 선택: 14일 우선, 없으면 1일 사용
        if '총 전환매출액(14일)' in df.columns:
            revenue_col = '총 전환매출액(14일)'
            data_type = '14일'
            logger.info('Using 14-day revenue data')
        elif '총 전환매출액(1일)' in df.columns:
            revenue_col = '총 전환매출액(1일)'
            data_type = '1일'
            warning_message = '⚠️ 주의: 14일 데이터가 없어 1일 데이터를 사용합니다. ROAS가 실제보다 낮게 표시될 수 있습니다.'
            logger.warning('Using 1-day revenue data (14-day not available)')
        else:
            logger.error('No revenue column found')
            return jsonify({'success': False, 'error': '매출액 컬럼 없음 (총 전환매출액(14일) 또는 총 전환매출액(1일) 필요)'}), 400

        # ROAS 컬럼 선택
        if '총광고수익률(14일)' in df.columns:
            roas_col = '총광고수익률(14일)'
        elif '총광고수익률(1일)' in df.columns:
            roas_col = '총광고수익률(1일)'
        else:
            roas_col = None

        # 주문수/판매수량 컬럼 선택
        if '총 주문수(14일)' in df.columns:
            order_col = '총 주문수(14일)'
            quantity_col = '총 판매수량(14일)' if '총 판매수량(14일)' in df.columns else '총 판매수량(1일)'
        else:
            order_col = '총 주문수(1일)'
            quantity_col = '총 판매수량(1일)'

        required_cols = required_cols_base + [order_col, quantity_col, revenue_col]
        if roas_col:
            required_cols.append(roas_col)

        missing = [col for col in required_cols if col not in df.columns]
        if missing:
            logger.error(f'Missing required columns: {missing}')
            return jsonify({'success': False, 'error': f'필수 컬럼 누락: {missing}'}), 400

        # 컬럼명 통일 (14일/1일 상관없이 동일한 이름으로 사용)
        df = df.rename(columns={
            revenue_col: '총 전환매출액',
            order_col: '총 주문수',
            quantity_col: '총 판매수량'
        })
        if roas_col:
            df = df.rename(columns={roas_col: '총광고수익률'})

        logger.info(f'Column mapping: revenue={revenue_col}, orders={order_col}')

        # 키워드 정규화: 연속된 공백을 하나로 통일
        if '키워드' in df.columns:
            original_keywords = df['키워드'].nunique()
            df['키워드'] = df['키워드'].astype(str).str.replace(r'\s+', ' ', regex=True).str.strip()
            normalized_keywords = df['키워드'].nunique()
            if original_keywords != normalized_keywords:
                logger.info(f'Keyword normalization: {original_keywords} → {normalized_keywords} unique keywords')

        # 데이터 정제
        # 1. 모든 광고 노출 지면 데이터 포함 (검색영역 + 비검색영역 + 리타겟팅)
        # 키워드가 '-'여도 포함 (비검색영역, 리타겟팅의 키워드는 '-'임)
        if '광고 노출 지면' in df.columns:
            exposure_types = df['광고 노출 지면'].value_counts()
            logger.info(f'Ad exposure types included: {exposure_types.to_dict()}')

        # 🔥 비검색영역 및 리타겟팅 통합 처리
        if '광고 노출 지면' in df.columns:
            # 1) 비검색영역 통합
            non_search_mask = df['광고 노출 지면'].str.contains('비검색', na=False)
            # 2) 리타겟팅 통합
            retargeting_mask = df['광고 노출 지면'].str.contains('리타겟팅', na=False)

            # 검색영역만 남김 (비검색, 리타겟팅 제외)
            search_only_df = df[~(non_search_mask | retargeting_mask)].copy()

            aggregated_rows = []

            # === 비검색영역 통합 ===
            if non_search_mask.sum() > 0:
                non_search_df = df[non_search_mask].copy()
                logger.info(f'비검색영역 통합 전: {non_search_mask.sum()}개 행')

                # 비검색영역 지표 합산
                non_search_aggregated = {
                    '키워드': '비검색영역 (통합)',
                    '광고 노출 지면': '비검색영역 (통합)',
                    '노출수': non_search_df['노출수'].sum(),
                    '클릭수': non_search_df['클릭수'].sum(),
                    '광고비': non_search_df['광고비'].sum(),
                    '총 주문수': non_search_df['총 주문수'].sum(),
                    '총 판매수량': non_search_df['총 판매수량'].sum(),
                    '총 전환매출액': non_search_df['총 전환매출액'].sum()
                }

                # 클릭률 재계산
                if non_search_aggregated['노출수'] > 0:
                    non_search_aggregated['클릭률'] = (non_search_aggregated['클릭수'] / non_search_aggregated['노출수']) * 100
                else:
                    non_search_aggregated['클릭률'] = 0

                # ROAS 재계산 (문자열 형식으로 저장하여 Excel 데이터와 일치)
                if non_search_aggregated['광고비'] > 0:
                    roas_value = (non_search_aggregated['총 전환매출액'] / non_search_aggregated['광고비']) * 100
                    non_search_aggregated['총광고수익률'] = f"{roas_value:.2f}%"
                else:
                    non_search_aggregated['총광고수익률'] = "0.00%"

                aggregated_rows.append(non_search_aggregated)
                logger.info(f'비검색영역 통합 완료: 1개 행으로 통합됨')
            else:
                logger.info('비검색영역 데이터 없음')

            # === 리타겟팅 통합 ===
            if retargeting_mask.sum() > 0:
                retargeting_df = df[retargeting_mask].copy()
                logger.info(f'리타겟팅 통합 전: {retargeting_mask.sum()}개 행')

                # 리타겟팅 지표 합산
                retargeting_aggregated = {
                    '키워드': '리타겟팅 (통합)',
                    '광고 노출 지면': '리타겟팅 (통합)',
                    '노출수': retargeting_df['노출수'].sum(),
                    '클릭수': retargeting_df['클릭수'].sum(),
                    '광고비': retargeting_df['광고비'].sum(),
                    '총 주문수': retargeting_df['총 주문수'].sum(),
                    '총 판매수량': retargeting_df['총 판매수량'].sum(),
                    '총 전환매출액': retargeting_df['총 전환매출액'].sum()
                }

                # 클릭률 재계산
                if retargeting_aggregated['노출수'] > 0:
                    retargeting_aggregated['클릭률'] = (retargeting_aggregated['클릭수'] / retargeting_aggregated['노출수']) * 100
                else:
                    retargeting_aggregated['클릭률'] = 0

                # ROAS 재계산 (문자열 형식으로 저장하여 Excel 데이터와 일치)
                if retargeting_aggregated['광고비'] > 0:
                    roas_value = (retargeting_aggregated['총 전환매출액'] / retargeting_aggregated['광고비']) * 100
                    retargeting_aggregated['총광고수익률'] = f"{roas_value:.2f}%"
                else:
                    retargeting_aggregated['총광고수익률'] = "0.00%"

                aggregated_rows.append(retargeting_aggregated)
                logger.info(f'리타겟팅 통합 완료: 1개 행으로 통합됨')
            else:
                logger.info('리타겟팅 데이터 없음')

            # 통합된 데이터 병합
            if aggregated_rows:
                aggregated_df = pd.DataFrame(aggregated_rows)
                df = pd.concat([search_only_df, aggregated_df], ignore_index=True)
            else:
                df = search_only_df

        logger.info(f'Total keywords to analyze (before dedup): {len(df)}개')

        # 🔥 키워드 중복 제거 - 동일 키워드는 데이터 합산
        if '키워드' in df.columns:
            keyword_groups = df.groupby('키워드', as_index=False).agg({
                '노출수': 'sum',
                '클릭수': 'sum',
                '광고비': 'sum',
                '총 주문수': 'sum',
                '총 판매수량': 'sum',
                '총 전환매출액': 'sum',
                '광고 노출 지면': 'first',  # 첫 번째 값 사용
            })

            # 클릭률 재계산 (Infinity 방지)
            keyword_groups['클릭률'] = (keyword_groups['클릭수'] / keyword_groups['노출수'] * 100).replace([np.inf, -np.inf], 0).fillna(0)

            # ROAS 재계산
            keyword_groups['총광고수익률'] = keyword_groups.apply(
                lambda row: f"{(row['총 전환매출액'] / row['광고비'] * 100):.2f}%" if row['광고비'] > 0 else "0.00%",
                axis=1
            )

            df = keyword_groups
            logger.info(f'Keyword deduplication completed: {len(df)}개 (unique keywords)')

        # 2. 클릭률 처리 (이미 % 형식이면 그대로, 소수점이면 100 곱하기)
        if df['클릭률'].max() <= 1:
            df['클릭률'] = df['클릭률'] * 100

        # 3. ROAS 파싱
        if df['총광고수익률'].dtype == 'object':
            # "356.78%" → 356.78 변환
            df['ROAS'] = df['총광고수익률'].str.rstrip('%').astype(float)
        else:
            df['ROAS'] = df['총광고수익률']
            if df['ROAS'].max() <= 10:  # 소수점 형식 (3.56 → 356)
                df['ROAS'] = df['ROAS'] * 100

        # 4. CPC 계산 (클릭당 단가) - Infinity 방지
        df['CPC'] = (df['광고비'] / df['클릭수']).replace([np.inf, -np.inf], 0).fillna(0)

        # 5. 결측치 및 Infinity 처리 (JSON 직렬화 오류 방지)
        numeric_cols = df.select_dtypes(include=[np.number]).columns
        for col in numeric_cols:
            df[col] = df[col].replace([np.inf, -np.inf], 0).fillna(0)

        logger.info(f'Processed {len(df)} valid keywords')

        # 요약 지표 계산
        total_spend = df['광고비'].sum()
        total_revenue = df['총 전환매출액'].sum()

        # 평균CTR 계산 (Infinity 방지)
        avg_ctr = df['클릭률'].mean()
        if np.isinf(avg_ctr) or np.isnan(avg_ctr):
            avg_ctr = 0

        summary = {
            '총광고비': int(total_spend),
            '총매출액': int(total_revenue),
            '평균ROAS': round((total_revenue / total_spend * 100), 2) if total_spend > 0 else 0,
            '총클릭수': int(df['클릭수'].sum()),
            '평균CTR': round(float(avg_ctr), 2),
            '총노출수': int(df['노출수'].sum()),
            '총주문수': int(df['총 주문수'].sum())
        }

        # JSON 안전 변환 함수
        import math
        def sanitize_for_json(obj):
            """Infinity, -Infinity, NaN을 JSON 안전 값으로 변환"""
            if isinstance(obj, (float, np.floating)):
                if math.isinf(float(obj)) or math.isnan(float(obj)):
                    return 0
                return float(obj)
            elif isinstance(obj, np.integer):
                return int(obj)
            return obj

        # DataFrame을 dict로 변환 후 Infinity/NaN 처리
        data = df.to_dict('records')
        for row in data:
            for key, value in row.items():
                row[key] = sanitize_for_json(value)

        # 세션에 저장 (선택사항)
        snapshot_id = int(pd.Timestamp.now().timestamp())
        session[f'coupang_snapshot_{snapshot_id}'] = {
            'data': data,
            'summary': summary,
            'created_at': pd.Timestamp.now().isoformat()
        }

        logger.info(f'Coupang data processed successfully: {len(data)} keywords')

        # JSON 응답 생성
        response_data = {
            'success': True,
            'data': data,
            'summary': summary,
            'data_type': data_type
        }

        # 경고 메시지가 있으면 포함
        if warning_message:
            response_data['warning'] = warning_message

        return jsonify(response_data)

    except Exception as e:
        logger.error(f'Coupang file upload failed: {e}')
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'error': f'처리 중 오류: {str(e)}'}), 500


@ad_bp.route('/api/ad-analysis/coupang-recommendations', methods=['POST'])
def coupang_recommendations():
    """
    쿠팡 광고 키워드 제외 추천 (향상된 0-100점 스코어링 시스템)

    Request Body:
        {
            "data": [...],  # 키워드 데이터
            "criteria": {
                "target_roas": 400  # 목표 ROAS (기본값: 400%)
            }
        }

    Response:
        {
            "success": true,
            "recommendations": [
                {
                    "keyword": "키워드명",
                    "score": 85,
                    "priority": "critical",
                    "reason": "즉시 제외 - ROAS 15%, 전환 0원",
                    "spend": 1200,
                    "revenue": 0,
                    "roas": 0,
                    "waste": 1200,
                    "waste_rate": 100,
                    "opportunity_loss": 4800,
                    "clicks": 10,
                    "ctr": 2.5,
                    "cpc": 120
                }
            ],
            "summary": {
                "total_waste": 39447,
                "total_opportunity_loss": 157788,
                "keywords_to_exclude": 182,
                "potential_savings": "94.0%",
                "critical_priority": 150,
                "high_priority": 32,
                "medium_priority": 0,
                "low_priority": 0
            }
        }
    """
    try:
        data = request.get_json()
        keywords = data.get('data', [])
        criteria = data.get('criteria', {})

        if not keywords:
            return jsonify({'success': False, 'error': '데이터가 없습니다'}), 400

        df = pd.DataFrame(keywords)
        logger.info(f'Analyzing {len(df)} keywords with enhanced scoring system')

        # ===== 중요: 검색영역만 추천 대상으로 분석 =====
        # 비검색영역, 리타겟팅은 추천에서 제외
        if '광고 노출 지면' in df.columns:
            original_count = len(df)
            df = df[df['광고 노출 지면'] == '검색 영역'].copy()
            logger.info(f'Filtered to search area only: {len(df)} keywords (from {original_count})')

        if len(df) == 0:
            return jsonify({
                'success': True,
                'recommendations': [],
                'summary': {
                    'total_waste': 0,
                    'total_opportunity_loss': 0,
                    'keywords_to_exclude': 0,
                    'potential_savings': '0%',
                    'critical_priority': 0,
                    'high_priority': 0,
                    'medium_priority': 0,
                    'low_priority': 0
                }
            })

        # 기본 통계 계산
        total_spend = df['광고비'].sum()
        total_revenue = df['총 전환매출액'].sum()
        avg_roas = (total_revenue / total_spend * 100) if total_spend > 0 else 0
        target_roas = criteria.get('target_roas', 400)  # 목표 ROAS 400%

        # === Phase 2: 중앙값 기반 통계 (Robust Statistics) ===
        median_cpc = df['CPC'].median()
        median_ctr = df['클릭률'].median()

        # CPC 백분위수
        cpc_percentiles = {
            'p25': df['CPC'].quantile(0.25),
            'p50': df['CPC'].quantile(0.50),
            'p75': df['CPC'].quantile(0.75),
            'p90': df['CPC'].quantile(0.90)
        }

        # 지출액 백분위수
        spend_percentiles = {
            'p25': df['광고비'].quantile(0.25),
            'p50': df['광고비'].quantile(0.50),
            'p75': df['광고비'].quantile(0.75),
            'p90': df['광고비'].quantile(0.90)
        }

        # 성과 구간별 통계
        tier_stats = {}
        tier_definitions = {
            'elite': df[df['ROAS'] >= 500],
            'high': df[(df['ROAS'] >= 300) & (df['ROAS'] < 500)],
            'mid': df[(df['ROAS'] >= 150) & (df['ROAS'] < 300)],
            'low': df[df['ROAS'] < 150]
        }

        for tier_name, tier_df in tier_definitions.items():
            if len(tier_df) > 0:
                tier_stats[tier_name] = {
                    'median_cpc': tier_df['CPC'].median(),
                    'p75_cpc': tier_df['CPC'].quantile(0.75),
                    'count': len(tier_df),
                    'avg_roas': tier_df['ROAS'].mean()
                }

        # 상위 성과 키워드 기준 (기회비용 계산용)
        top_performers = df[df['ROAS'] >= target_roas]
        if len(top_performers) > 0:
            top_avg_roas = top_performers['ROAS'].mean()
        else:
            top_avg_roas = avg_roas

        recommendations = []

        for _, row in df.iterrows():
            keyword = row['키워드']
            spend = float(row['광고비'])
            revenue = float(row['총 전환매출액'])
            roas = float(row.get('ROAS', 0))
            clicks = int(row['클릭수'])
            ctr = float(row['클릭률'])
            cpc = float(row['CPC'])

            # === Phase 2: 새로운 스코어링 시스템 ===
            reasons = []

            # === 1. 수익성 점수 (0-50점) - ROAS 기반 ===
            if revenue == 0:
                profitability_score = 50
                reasons.append("전환 0원")
            elif roas < 20:
                profitability_score = 45
                reasons.append(f"ROAS {roas:.1f}% (극심한 손실)")
            elif roas < 50:
                profitability_score = 40
                reasons.append(f"ROAS {roas:.1f}% (심각한 손실)")
            elif roas < 100:
                profitability_score = 35
                reasons.append(f"ROAS {roas:.1f}% (손실)")
            elif roas < 150:
                profitability_score = 25
                reasons.append(f"ROAS {roas:.1f}% (낮은 수익)")
            elif roas < 200:
                profitability_score = 15
                reasons.append(f"ROAS {roas:.1f}% (목표 미달)")
            elif roas < 300:
                profitability_score = 10
                reasons.append(f"ROAS {roas:.1f}% (목표 근접)")
            else:
                profitability_score = 0  # ROAS >= 300%

            # === 2. 효율성 점수 (0-25점) - 성과 구간별 CPC 비교 ===
            # 키워드 성과 구간 판정
            if roas >= 500:
                tier = 'elite'
            elif roas >= 300:
                tier = 'high'
            elif roas >= 150:
                tier = 'mid'
            else:
                tier = 'low'

            # 해당 구간의 중앙값 CPC
            if tier in tier_stats and tier_stats[tier]['count'] >= 3:
                tier_median_cpc = tier_stats[tier]['median_cpc']
            else:
                tier_median_cpc = median_cpc  # fallback

            # CPC 비율 계산
            if tier_median_cpc > 0:
                cpc_ratio = cpc / tier_median_cpc
            else:
                cpc_ratio = 1.0

            # 성과 구간별로 다른 기준 적용
            if tier in ['elite', 'high']:
                # 고성과 키워드: CPC 기준 관대
                if cpc_ratio > 3.0:
                    efficiency_score = 10
                    reasons.append(f"CPC 과다 ({cpc:.0f}원)")
                elif cpc_ratio > 2.5:
                    efficiency_score = 5
                else:
                    efficiency_score = 0
            elif tier == 'mid':
                # 중성과 키워드: 보통 기준
                if cpc_ratio > 2.5:
                    efficiency_score = 20
                    reasons.append(f"CPC 높음 ({cpc:.0f}원)")
                elif cpc_ratio > 2.0:
                    efficiency_score = 15
                elif cpc_ratio > 1.5:
                    efficiency_score = 10
                else:
                    efficiency_score = 0
            else:
                # 저성과 키워드: CPC 기준 엄격
                if cpc_ratio > 2.0:
                    efficiency_score = 25
                    reasons.append(f"CPC 과다 ({cpc:.0f}원)")
                elif cpc_ratio > 1.5:
                    efficiency_score = 20
                elif cpc_ratio > 1.2:
                    efficiency_score = 15
                else:
                    efficiency_score = 5

            # === 3. 규모 리스크 점수 (0-25점) - 지출액 + ROAS 조합 ===
            # 지출 수준 판정
            if spend > spend_percentiles['p90']:
                spend_level = 'very_high'
            elif spend > spend_percentiles['p75']:
                spend_level = 'high'
            elif spend > spend_percentiles['p50']:
                spend_level = 'medium'
            else:
                spend_level = 'low'

            # ROAS와 지출 조합으로 점수 계산
            if roas == 0:
                # 전환 0원 케이스
                if spend_level == 'very_high':
                    scale_risk_score = 25
                    reasons.append(f"고지출 ({spend:,.0f}원)")
                elif spend_level == 'high':
                    scale_risk_score = 20
                    reasons.append(f"중간 지출")
                elif spend_level == 'medium':
                    scale_risk_score = 15
                else:
                    scale_risk_score = 10
            elif roas < 100:
                # 손실 케이스
                if spend_level == 'very_high':
                    scale_risk_score = 20
                    reasons.append(f"고지출 ({spend:,.0f}원)")
                elif spend_level == 'high':
                    scale_risk_score = 15
                elif spend_level == 'medium':
                    scale_risk_score = 10
                else:
                    scale_risk_score = 5
            elif roas < 200:
                # 낮은 수익 케이스
                if spend_level in ['very_high', 'high']:
                    scale_risk_score = 10
                else:
                    scale_risk_score = 0
            elif roas < 300:
                # 목표 미달 케이스
                if spend_level == 'very_high':
                    scale_risk_score = 5
                else:
                    scale_risk_score = 0
            else:
                # 목표 달성 (ROAS >= 300%)
                scale_risk_score = 0

            # 총점 계산
            score = profitability_score + efficiency_score + scale_risk_score

            # === Phase 3: 개선된 우선순위 결정 (조건 기반) ===
            # 0단계: 데이터 부족 판정
            if spend < EXCLUDE_MIN_SPEND and clicks < EXCLUDE_MIN_CLICKS:
                priority = None
                priority_label = '데이터 부족'
                reasons.append('데이터 부족')
            # 1단계: 고CPC + 저ROAS → 즉시제외 (광고비 무관)
            elif cpc >= EXCLUDE_CPC_CRITICAL and roas < 100:
                priority = 'critical'
                priority_label = '즉시 제외'
                reasons.append(f'고CPC({cpc:.0f}원) + 저ROAS')
            elif cpc >= EXCLUDE_CPC_VERY_HIGH and roas < 200:
                priority = 'critical'
                priority_label = '즉시 제외'
                reasons.append(f'초고CPC({cpc:.0f}원) + 저ROAS')
            # 2단계: 전환없음 (ROAS 0%)
            elif revenue == 0:
                if clicks >= EXCLUDE_CLICKS_CRITICAL:
                    priority = 'critical'
                    priority_label = '즉시 제외'
                    reasons.append(f'{clicks}클릭 전환없음')
                elif clicks >= EXCLUDE_CLICKS_HIGH:
                    priority = 'high'
                    priority_label = '조속히 제외'
                    reasons.append(f'{clicks}클릭 전환없음')
                else:
                    priority = 'medium'
                    priority_label = '검토 필요'
                    reasons.append('전환없음 검토')
            # 3단계: ROAS 1~100% (손실)
            elif roas < 100:
                if spend >= spend_percentiles['p75']:
                    priority = 'critical'
                    priority_label = '즉시 제외'
                    reasons.append('저ROAS + 고지출')
                elif spend >= spend_percentiles['p50']:
                    priority = 'high'
                    priority_label = '조속히 제외'
                    reasons.append('저ROAS + 중지출')
                else:
                    priority = 'medium'
                    priority_label = '검토 필요'
            # 4단계: ROAS 100~200% (저조)
            elif roas < 200:
                if spend >= spend_percentiles['p75']:
                    priority = 'high'
                    priority_label = '조속히 제외'
                    reasons.append('저조ROAS + 고지출')
                else:
                    priority = 'medium'
                    priority_label = '검토 필요'
            # 5단계: ROAS 200~300% (목표 근접)
            elif roas < 300:
                priority = 'medium'
                priority_label = '검토 필요'
                reasons.append('ROAS 개선필요')
            # 6단계: ROAS 300%+ (양호)
            else:
                priority = 'low'
                priority_label = '모니터링'

            # === 낭비 및 기회비용 계산 ===
            if spend == 0:
                # 광고비 0원인 경우 - 비정상 데이터 (낭비 없음)
                waste = 0
                waste_rate = 0
                expected_revenue = 0
                opportunity_loss = 0
            elif roas < 100:
                # 손실 케이스: 광고비 - 매출
                waste = spend - revenue
                waste_rate = 100 - roas
                # 기회비용: 이 광고비를 상위 성과 키워드에 투자했을 때의 기대 매출
                expected_revenue = spend * (top_avg_roas / 100)
                opportunity_loss = expected_revenue - revenue
            else:
                # 목표 미달 케이스: 낭비 없음
                waste = 0
                waste_rate = 0
                # 기회비용: 이 광고비를 상위 성과 키워드에 투자했을 때의 기대 매출
                expected_revenue = spend * (top_avg_roas / 100)
                opportunity_loss = expected_revenue - revenue

            # === 추천 사유 생성 ===
            reason = f"{priority_label} - " + ", ".join(reasons[:3])  # 최대 3개 사유

            recommendations.append({
                'keyword': keyword,
                'score': int(score),
                'priority': priority,
                'reason': reason,
                'spend': spend,
                'revenue': revenue,
                'roas': roas,
                'waste': float(waste),
                'waste_rate': float(waste_rate),
                'opportunity_loss': float(opportunity_loss),
                'clicks': clicks,
                'ctr': ctr,
                'cpc': cpc
            })

        # === 정렬: 점수 높은 순 ===
        recommendations.sort(key=lambda x: -x['score'])

        # === 요약 통계 ===
        total_waste = sum(r['waste'] for r in recommendations)
        total_opportunity_loss = sum(r['opportunity_loss'] for r in recommendations)

        # 데이터 부족 키워드 분리
        insufficient_data = [r for r in recommendations if r['priority'] is None]
        valid_recommendations = [r for r in recommendations if r['priority'] is not None]

        summary = {
            'total_waste': int(total_waste),
            'total_opportunity_loss': int(total_opportunity_loss),
            'keywords_to_exclude': len(valid_recommendations),
            'potential_savings': f"{(total_waste / total_spend * 100):.1f}%" if total_spend > 0 else "0%",
            'critical_priority': len([r for r in valid_recommendations if r['priority'] == 'critical']),
            'high_priority': len([r for r in valid_recommendations if r['priority'] == 'high']),
            'medium_priority': len([r for r in valid_recommendations if r['priority'] == 'medium']),
            'low_priority': len([r for r in valid_recommendations if r['priority'] == 'low']),
            'insufficient_data': len(insufficient_data),
            'avg_score': int(sum(r['score'] for r in valid_recommendations) / len(valid_recommendations)) if valid_recommendations else 0
        }

        logger.info(f'Generated {len(recommendations)} recommendations (avg score: {summary["avg_score"]}, total waste: {total_waste:.0f}원)')

        return jsonify({
            'success': True,
            'recommendations': recommendations,
            'summary': summary
        })

    except Exception as e:
        logger.error(f'Recommendation generation failed: {e}')
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'error': f'추천 생성 실패: {str(e)}'}), 500


@ad_bp.route('/api/ad-analysis/manual-input', methods=['POST'])
def manual_input():
    """수기 데이터 입력 (데이터베이스 저장)"""
    user_id = get_current_user_id()

    try:
        data = request.get_json()

        if not data or 'data' not in data:
            return jsonify({'success': False, 'error': '데이터가 없습니다'}), 400

        # DataFrame 생성
        df = pd.DataFrame(data['data'])

        # 필수 컬럼 확인
        required_cols = ['date', 'campaign_name', 'spend', 'clicks', 'conversions', 'revenue']
        missing_cols = [col for col in required_cols if col not in df.columns]

        if missing_cols:
            return jsonify({
                'success': False,
                'error': f'필수 필드가 누락되었습니다: {", ".join(missing_cols)}'
            }), 400

        snapshot_name = data.get('snapshot_name', f'수기입력 {pd.Timestamp.now().strftime("%Y-%m-%d %H:%M")}')

        # 임시 스냅샷 ID 생성 (DB 대신 메모리 사용)
        snapshot_id = int(pd.Timestamp.now().timestamp())

        # In-Memory 방식으로 지표 계산 (DB 없이)
        metrics = _calculate_metrics_inmemory(df)

        # 세션에 저장 (선택사항)
        session[f'snapshot_{snapshot_id}'] = {
            'name': snapshot_name,
            'data': df.to_dict('records'),
            'metrics': metrics,
            'created_at': pd.Timestamp.now().isoformat()
        }

        logger.info(f'Manual data input processed in-memory: {len(df)} rows, snapshot_id: {snapshot_id}')

        return jsonify({
            'success': True,
            'snapshot_id': snapshot_id,
            'metrics': metrics
        })

    except Exception as e:
        logger.error(f'Manual input failed: {e}')
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'error': f'처리 중 오류: {str(e)}'}), 500


# ========================================
# 3. 분석 관리 API
# ========================================

@ad_bp.route('/api/ad-analysis/snapshots')
def get_snapshots():
    """
    저장된 분석 목록 조회

    Query Params:
        - saved_only: true/false (저장된 것만)

    Response:
        {
            "snapshots": [...]
        }
    """
    user_id = get_current_user_id()  # 테스트용 임시 user_id
    saved_only = request.args.get('saved_only', 'false').lower() == 'true'

    try:
        analyzer = AdAnalyzer(user_id)
        snapshots = analyzer.get_snapshots(saved_only)

        return jsonify({'snapshots': snapshots})

    except Exception as e:
        logger.error(f"Get snapshots failed: {e}")
        return create_error_response("분석 목록 조회 실패", 500)


@ad_bp.route('/api/ad-analysis/snapshots/<int:snapshot_id>')
def get_snapshot_detail(snapshot_id):
    """
    특정 분석 상세 조회

    Response:
        {
            "snapshot": {...},
            "daily_data": [...],
            "metrics": {...},
            "insights": "...",
            "campaigns": [...]
        }
    """
    user_id = get_current_user_id()  # 테스트용 임시 user_id

    try:
        analyzer = AdAnalyzer(user_id)

        # 소유권 확인
        if not analyzer.check_ownership(snapshot_id):
            return create_error_response("접근 권한이 없습니다", 403)

        data = analyzer.get_snapshot_detail(snapshot_id)

        return jsonify(data)

    except ValueError as e:
        return create_error_response(str(e), 404)

    except Exception as e:
        logger.error(f"Get snapshot detail failed: {e}")
        return create_error_response("분석 조회 실패", 500)


@ad_bp.route('/api/ad-analysis/snapshots/<int:snapshot_id>', methods=['PUT'])
def update_snapshot(snapshot_id):
    """
    분석 저장/수정

    Request Body:
        {
            "is_saved": true,
            "snapshot_name": "수정된 이름",
            "tags": "블프,신규",
            "memo": "메모 내용"
        }

    Response:
        {"success": true}
    """
    user_id = get_current_user_id()  # 테스트용 임시 user_id

    try:
        data = request.get_json()

        analyzer = AdAnalyzer(user_id)

        # 소유권 확인
        if not analyzer.check_ownership(snapshot_id):
            return create_error_response("접근 권한이 없습니다", 403)

        success = analyzer.update_snapshot(snapshot_id, data)

        if success:
            return jsonify(create_success_response(message="수정되었습니다"))
        else:
            return create_error_response("수정 실패", 500)

    except Exception as e:
        logger.error(f"Update snapshot failed: {e}")
        return create_error_response("수정 중 오류가 발생했습니다", 500)


@ad_bp.route('/api/ad-analysis/snapshots/<int:snapshot_id>', methods=['DELETE'])
def delete_snapshot(snapshot_id):
    """
    분석 삭제

    Response:
        {"success": true}
    """
    user_id = get_current_user_id()  # 테스트용 임시 user_id

    try:
        analyzer = AdAnalyzer(user_id)

        # 소유권 확인
        if not analyzer.check_ownership(snapshot_id):
            return create_error_response("접근 권한이 없습니다", 403)

        success = analyzer.delete_snapshot(snapshot_id)

        if success:
            return jsonify(create_success_response(message="삭제되었습니다"))
        else:
            return create_error_response("삭제 실패", 404)

    except Exception as e:
        logger.error(f"Delete snapshot failed: {e}")
        return create_error_response("삭제 중 오류가 발생했습니다", 500)


# ========================================
# 4. 비교 분석 API
# ========================================

@ad_bp.route('/api/ad-analysis/compare')
def compare_periods():
    """
    기간 비교 분석

    Query Params:
        - snapshot_a: 기준 분석 ID
        - snapshot_b: 비교 분석 ID

    Response:
        {
            "comparison": {...},
            "summary": "..."
        }
    """
    user_id = get_current_user_id()  # 테스트용 임시 user_id

    snapshot_a = request.args.get('snapshot_a', type=int)
    snapshot_b = request.args.get('snapshot_b', type=int)

    if not snapshot_a or not snapshot_b:
        return create_error_response("두 개의 스냅샷 ID가 필요합니다", 400)

    try:
        analyzer = AdAnalyzer(user_id)

        # 소유권 확인
        if not analyzer.check_ownership(snapshot_a) or not analyzer.check_ownership(snapshot_b):
            return create_error_response("접근 권한이 없습니다", 403)

        comparison = analyzer.compare_snapshots(snapshot_a, snapshot_b)

        return jsonify(comparison)

    except Exception as e:
        logger.error(f"Compare snapshots failed: {e}")
        return create_error_response("비교 분석 실패", 500)


# ========================================
# 5. 목표 관리 API
# ========================================

@ad_bp.route('/api/ad-analysis/goals', methods=['GET', 'POST'])
def manage_goals():
    """
    월별 목표 설정/조회

    GET - Query Params:
        - year_month: YYYY-MM

    POST - Request Body:
        {
            "year_month": "2024-11",
            "budget": 10000000,
            "target_roas": 4.0,
            "target_revenue": 40000000
        }

    Response:
        {"goal": {...}}  (GET)
        {"success": true}  (POST)
    """
    user_id = get_current_user_id()  # 테스트용 임시 user_id

    if request.method == 'GET':
        year_month = request.args.get('year_month')

        if not year_month:
            return create_error_response("year_month 파라미터가 필요합니다", 400)

        try:
            sql = """
                SELECT * FROM ad_monthly_goals
                WHERE user_id = %s AND year_month = %s
            """
            goal = execute_query(sql, (user_id, year_month), fetch_one=True)

            return jsonify({'goal': goal})

        except Exception as e:
            logger.error(f"Get goal failed: {e}")
            return create_error_response("목표 조회 실패", 500)

    else:  # POST
        try:
            data = request.get_json()

            year_month = data.get('year_month')
            budget = data.get('budget')
            target_roas = data.get('target_roas')
            target_revenue = data.get('target_revenue')

            if not year_month:
                return create_error_response("year_month가 필요합니다", 400)

            # UPSERT (ON DUPLICATE KEY UPDATE)
            sql = """
                INSERT INTO ad_monthly_goals
                (user_id, year_month, budget, target_roas, target_revenue)
                VALUES (%s, %s, %s, %s, %s)
                ON DUPLICATE KEY UPDATE
                    budget = VALUES(budget),
                    target_roas = VALUES(target_roas),
                    target_revenue = VALUES(target_revenue)
            """

            execute_insert(sql, (user_id, year_month, budget, target_roas, target_revenue))

            return jsonify(create_success_response(message="목표가 저장되었습니다"))

        except Exception as e:
            logger.error(f"Save goal failed: {e}")
            return create_error_response("목표 저장 실패", 500)


@ad_bp.route('/api/ad-analysis/budget-pacing')
def budget_pacing():
    """
    예산 소진율 및 페이싱 분석

    Query Params:
        - year_month: YYYY-MM

    Response:
        {
            "budget": 10000000,
            "spent": 5800000,
            "spent_rate": 58,
            "status": "FAST",
            ...
        }
    """
    user_id = get_current_user_id()  # 테스트용 임시 user_id
    year_month = request.args.get('year_month')

    if not year_month:
        return create_error_response("year_month 파라미터가 필요합니다", 400)

    try:
        analyzer = AdAnalyzer(user_id)
        pacing = analyzer.calculate_budget_pacing(year_month)

        return jsonify(pacing)

    except Exception as e:
        logger.error(f"Budget pacing failed: {e}")
        return create_error_response("예산 분석 실패", 500)


# ========================================
# 6. 캠페인 메모 API
# ========================================

@ad_bp.route('/api/ad-analysis/memos', methods=['GET', 'POST'])
def manage_memos():
    """
    캠페인 메모 조회/추가

    GET - Query Params:
        - campaign_name: 캠페인명

    POST - Request Body:
        {
            "campaign_name": "블프_신규",
            "memo": "소재 #3으로 교체"
        }

    Response:
        {"memos": [...]}  (GET)
        {"success": true}  (POST)
    """
    user_id = get_current_user_id()  # 테스트용 임시 user_id

    if request.method == 'GET':
        campaign_name = request.args.get('campaign_name')

        try:
            sql = """
                SELECT * FROM ad_campaign_memos
                WHERE user_id = %s
            """
            params = [user_id]

            if campaign_name:
                sql += " AND campaign_name = %s"
                params.append(campaign_name)

            sql += " ORDER BY created_at DESC"

            memos = execute_query(sql, tuple(params))

            # 날짜 포맷팅
            for memo in memos:
                memo['created_at'] = memo['created_at'].strftime('%Y-%m-%d %H:%M:%S')

            return jsonify({'memos': memos})

        except Exception as e:
            logger.error(f"Get memos failed: {e}")
            return create_error_response("메모 조회 실패", 500)

    else:  # POST
        try:
            data = request.get_json()

            campaign_name = data.get('campaign_name')
            memo = data.get('memo')

            if not campaign_name or not memo:
                return create_error_response("campaign_name과 memo가 필요합니다", 400)

            sql = """
                INSERT INTO ad_campaign_memos
                (user_id, campaign_name, memo)
                VALUES (%s, %s, %s)
            """

            execute_insert(sql, (user_id, campaign_name, memo))

            return jsonify(create_success_response(message="메모가 저장되었습니다"))

        except Exception as e:
            logger.error(f"Save memo failed: {e}")
            return create_error_response("메모 저장 실패", 500)


# ========================================
# 7. 리포트 내보내기 API
# ========================================

@ad_bp.route('/api/ad-analysis/export/pdf/<int:snapshot_id>')
def export_pdf(snapshot_id):
    """
    PDF 리포트 생성 및 다운로드
    """
    user_id = get_current_user_id()  # 테스트용 임시 user_id

    try:
        analyzer = AdAnalyzer(user_id)

        if not analyzer.check_ownership(snapshot_id):
            return create_error_response("접근 권한이 없습니다", 403)

        # TODO: PDF 생성 로직 구현
        return create_error_response("PDF 내보내기 기능은 준비 중입니다", 501)

    except Exception as e:
        logger.error(f"Export PDF failed: {e}")
        return create_error_response("PDF 생성 실패", 500)


@ad_bp.route('/api/ad-analysis/export/excel/<int:snapshot_id>')
def export_excel(snapshot_id):
    """
    Excel 리포트 생성 및 다운로드
    """
    user_id = get_current_user_id()  # 테스트용 임시 user_id

    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils.dataframe import dataframe_to_rows
        from datetime import datetime
        import tempfile

        analyzer = AdAnalyzer(user_id)

        if not analyzer.check_ownership(snapshot_id):
            return create_error_response("접근 권한이 없습니다", 403)

        # 데이터 조회
        snapshot_data = analyzer.get_snapshot_detail(snapshot_id)
        if not snapshot_data:
            return create_error_response("분석 데이터를 찾을 수 없습니다", 404)

        metrics = snapshot_data.get('metrics', {})
        snapshot_info = snapshot_data.get('snapshot', {})

        # Excel 워크북 생성
        wb = Workbook()

        # 스타일 정의
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True, size=12)
        header_alignment = Alignment(horizontal="center", vertical="center")
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # Sheet 1: 요약
        ws_summary = wb.active
        ws_summary.title = "요약"

        # 제목
        ws_summary['A1'] = "광고 분석 리포트"
        ws_summary['A1'].font = Font(size=16, bold=True)
        ws_summary.merge_cells('A1:D1')

        # 기본 정보
        ws_summary['A3'] = "분석명"
        ws_summary['B3'] = snapshot_info.get('snapshot_name', 'N/A')
        ws_summary['A4'] = "분석 기간"
        ws_summary['B4'] = f"{snapshot_info.get('period_start', 'N/A')} ~ {snapshot_info.get('period_end', 'N/A')}"
        ws_summary['A5'] = "생성일시"
        ws_summary['B5'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')

        # 주요 지표
        ws_summary['A7'] = "주요 지표"
        ws_summary['A7'].font = Font(size=14, bold=True)

        summary_data = [
            ["지표", "값"],
            ["총 지출", f"{metrics.get('total_spend', 0):,.0f}원"],
            ["총 매출", f"{metrics.get('total_revenue', 0):,.0f}원"],
            ["평균 ROAS", f"{metrics.get('avg_roas', 0):.2f}"],
            ["평균 CTR", f"{metrics.get('avg_ctr', 0):.2f}%"],
            ["평균 CPC", f"{metrics.get('avg_cpc', 0):,.0f}원"],
            ["평균 CPA", f"{metrics.get('avg_cpa', 0):,.0f}원"],
            ["전환율", f"{metrics.get('cvr', 0):.2f}%"],
            ["총 클릭", f"{metrics.get('total_clicks', 0):,}"],
            ["총 전환", f"{metrics.get('total_conversions', 0):,}"],
        ]

        for row_idx, row_data in enumerate(summary_data, start=8):
            for col_idx, value in enumerate(row_data, start=1):
                cell = ws_summary.cell(row=row_idx, column=col_idx, value=value)
                if row_idx == 8:  # 헤더
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = header_alignment
                cell.border = border

        # 열 너비 조정
        ws_summary.column_dimensions['A'].width = 20
        ws_summary.column_dimensions['B'].width = 25

        # Sheet 2: 캠페인 성과
        ws_campaigns = wb.create_sheet("캠페인 성과")
        campaigns = metrics.get('campaigns', [])

        if campaigns:
            campaign_headers = ["순위", "캠페인명", "광고유형", "ROAS", "CTR(%)", "CPA(원)", "CVR(%)", "지출(원)", "매출(원)", "클릭", "전환"]

            # 헤더 작성
            for col_idx, header in enumerate(campaign_headers, start=1):
                cell = ws_campaigns.cell(row=1, column=col_idx, value=header)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = header_alignment
                cell.border = border

            # 데이터 작성
            for row_idx, campaign in enumerate(campaigns, start=2):
                row_data = [
                    campaign.get('rank', row_idx - 1),
                    campaign.get('campaign_name', 'N/A'),
                    "매출형" if campaign.get('ad_type') == 'sales' else "잠재고객",
                    round(campaign.get('roas', 0), 2),
                    round(campaign.get('ctr', 0), 2),
                    int(campaign.get('cpa', 0)),
                    round(campaign.get('cvr', 0), 2),
                    int(campaign.get('spend', 0)),
                    int(campaign.get('revenue', 0)),
                    int(campaign.get('clicks', 0)),
                    int(campaign.get('conversions', 0))
                ]

                for col_idx, value in enumerate(row_data, start=1):
                    cell = ws_campaigns.cell(row=row_idx, column=col_idx, value=value)
                    cell.border = border

                    # 숫자 서식
                    if col_idx in [4, 5, 7]:  # ROAS, CTR, CVR
                        cell.number_format = '0.00'
                    elif col_idx in [6, 8, 9]:  # CPA, 지출, 매출
                        cell.number_format = '#,##0'

            # 열 너비 자동 조정
            for col in ws_campaigns.columns:
                max_length = 0
                column = col[0].column_letter
                for cell in col:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws_campaigns.column_dimensions[column].width = adjusted_width

        # Sheet 3: 일별 데이터
        ws_daily = wb.create_sheet("일별 데이터")
        daily_data = metrics.get('daily_data', metrics.get('daily_trend', []))

        if daily_data:
            # 헤더 결정 (campaign_name 포함 여부)
            sample_row = daily_data[0] if daily_data else {}
            has_campaign = 'campaign_name' in sample_row

            if has_campaign:
                daily_headers = ["날짜", "캠페인명", "지출(원)", "매출(원)", "ROAS", "클릭", "전환", "CTR(%)", "CVR(%)"]
            else:
                daily_headers = ["날짜", "지출(원)", "매출(원)", "ROAS", "클릭", "전환", "CTR(%)", "CVR(%)"]

            # 헤더 작성
            for col_idx, header in enumerate(daily_headers, start=1):
                cell = ws_daily.cell(row=1, column=col_idx, value=header)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = header_alignment
                cell.border = border

            # 데이터 작성
            for row_idx, daily in enumerate(daily_data, start=2):
                if has_campaign:
                    row_data = [
                        daily.get('date', 'N/A'),
                        daily.get('campaign_name', 'N/A'),
                        int(daily.get('spend', 0)),
                        int(daily.get('revenue', 0)),
                        round(daily.get('roas', 0), 2),
                        int(daily.get('clicks', 0)),
                        int(daily.get('conversions', 0)),
                        round(daily.get('ctr', 0), 2),
                        round(daily.get('cvr', 0), 2)
                    ]
                else:
                    row_data = [
                        daily.get('date', 'N/A'),
                        int(daily.get('spend', 0)),
                        int(daily.get('revenue', 0)),
                        round(daily.get('roas', 0), 2),
                        int(daily.get('clicks', 0)),
                        int(daily.get('conversions', 0)),
                        round(daily.get('ctr', 0), 2),
                        round(daily.get('cvr', 0), 2)
                    ]

                for col_idx, value in enumerate(row_data, start=1):
                    cell = ws_daily.cell(row=row_idx, column=col_idx, value=value)
                    cell.border = border

            # 열 너비 조정
            for col in ws_daily.columns:
                max_length = 0
                column = col[0].column_letter
                for cell in col:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws_daily.column_dimensions[column].width = adjusted_width

        # Sheet 4: 소재 성과 (있는 경우)
        creatives = metrics.get('creatives', [])
        if creatives:
            ws_creatives = wb.create_sheet("소재 성과")

            creative_headers = ["순위", "소재명", "플랫폼", "유형", "ROAS", "CTR(%)", "CVR(%)", "지출(원)", "매출(원)", "클릭", "전환"]

            # 헤더
            for col_idx, header in enumerate(creative_headers, start=1):
                cell = ws_creatives.cell(row=1, column=col_idx, value=header)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = header_alignment
                cell.border = border

            # 데이터
            for row_idx, creative in enumerate(creatives, start=2):
                row_data = [
                    creative.get('roas_rank', row_idx - 1),
                    creative.get('ad_creative_name', 'N/A'),
                    creative.get('platform', 'N/A'),
                    creative.get('creative_type', 'N/A'),
                    round(creative.get('roas', 0), 2),
                    round(creative.get('ctr', 0), 2),
                    round(creative.get('cvr', 0), 2),
                    int(creative.get('spend', 0)),
                    int(creative.get('revenue', 0)),
                    int(creative.get('clicks', 0)),
                    int(creative.get('conversions', 0))
                ]

                for col_idx, value in enumerate(row_data, start=1):
                    cell = ws_creatives.cell(row=row_idx, column=col_idx, value=value)
                    cell.border = border

            # 열 너비 조정
            for col in ws_creatives.columns:
                max_length = 0
                column = col[0].column_letter
                for cell in col:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws_creatives.column_dimensions[column].width = adjusted_width

        # 임시 파일로 저장
        temp_dir = tempfile.gettempdir()
        filename = f"ad_report_{snapshot_id}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        filepath = os.path.join(temp_dir, filename)

        wb.save(filepath)

        logger.info(f"Excel file created: {filepath}")

        # 파일 전송
        return send_file(
            filepath,
            as_attachment=True,
            download_name=filename,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )

    except ImportError:
        logger.error("openpyxl not installed")
        return create_error_response("Excel 라이브러리가 설치되지 않았습니다. pip install openpyxl을 실행하세요.", 500)
    except Exception as e:
        logger.error(f"Export Excel failed: {e}")
        import traceback
        traceback.print_exc()
        return create_error_response(f"Excel 생성 실패: {str(e)}", 500)


# ========================================
# 8. 템플릿 다운로드 API
# ========================================

@ad_bp.route('/api/ad-analysis/template/<template_type>')
def download_template(template_type):
    """
    Excel 템플릿 다운로드

    template_type: 'unified', 'naver', 'meta', 'google', 'kakao', 'generic'
    """
    templates = {
        'unified': 'ad_template_unified.xlsx',  # 통합 템플릿 (한글 컬럼, 광고유형 포함)
        'generic': 'ad_template_generic.xlsx',
        'naver': 'ad_template_naver.xlsx',
        'meta': 'ad_template_meta.xlsx',
        'google': 'ad_template_google.xlsx',
        'kakao': 'ad_template_kakao.xlsx'
    }

    filename = templates.get(template_type, templates['generic'])
    template_path = os.path.join(current_app.root_path, 'static', 'templates', filename)

    if not os.path.exists(template_path):
        return create_error_response("템플릿 파일을 찾을 수 없습니다", 404)

    return send_file(template_path, as_attachment=True, download_name=filename)


# ========================================
# Helper Functions
# ========================================

def _calculate_creative_metrics(df):
    """
    소재별 성과 지표 계산

    Args:
        df: pandas DataFrame with columns including ad_creative_name

    Returns:
        list: 소재별 지표 리스트 (ROAS 순위순)
    """
    import numpy as np

    # ad_creative_name으로 그룹화
    creative_stats = df.groupby('ad_creative_name').agg({
        'spend': 'sum',
        'revenue': 'sum',
        'clicks': 'sum',
        'conversions': 'sum',
        'impressions': 'sum' if 'impressions' in df.columns else lambda x: 0
    }).reset_index()

    # ROAS 계산
    creative_stats['roas'] = (creative_stats['revenue'] / creative_stats['spend']).round(2)

    # CTR 계산
    if 'impressions' in df.columns and creative_stats['impressions'].sum() > 0:
        creative_stats['ctr'] = (creative_stats['clicks'] / creative_stats['impressions'] * 100).round(2)
    else:
        creative_stats['ctr'] = 0

    # CPA 계산
    creative_stats['cpa'] = (creative_stats['spend'] / creative_stats['conversions']).replace([np.inf, -np.inf], 0).round(0)

    # CVR (전환율) 계산
    creative_stats['cvr'] = (creative_stats['conversions'] / creative_stats['clicks'] * 100).replace([np.inf, -np.inf], 0).round(2)

    # CPC 계산
    creative_stats['cpc'] = (creative_stats['spend'] / creative_stats['clicks']).replace([np.inf, -np.inf], 0).round(0)

    # 객단가 계산
    creative_stats['avg_order_value'] = (creative_stats['revenue'] / creative_stats['conversions']).replace([np.inf, -np.inf], 0).round(0)

    # ROAS 순위 계산
    creative_stats = creative_stats.sort_values('roas', ascending=False)
    creative_stats['roas_rank'] = range(1, len(creative_stats) + 1)

    # CVR 순위 계산 (별도 정렬)
    creative_stats_cvr_sorted = creative_stats.sort_values('cvr', ascending=False)
    creative_stats['cvr_rank'] = creative_stats_cvr_sorted.index.map(lambda x: creative_stats_cvr_sorted.index.get_loc(x) + 1)

    # 소재 타입 정보 추가 (있는 경우)
    if 'ad_creative_type' in df.columns:
        creative_type_map = df.groupby('ad_creative_name')['ad_creative_type'].first()
        creative_stats['creative_type'] = creative_stats['ad_creative_name'].map(creative_type_map)

    # 플랫폼 정보 추가 (있는 경우)
    if 'platform' in df.columns:
        platform_map = df.groupby('ad_creative_name')['platform'].first()
        creative_stats['platform'] = creative_stats['ad_creative_name'].map(platform_map)

    # 상태 판정 (ROAS 기준)
    def get_creative_status(roas):
        if roas >= 4.0:
            return 'excellent'
        elif roas >= 3.0:
            return 'good'
        else:
            return 'poor'

    creative_stats['status'] = creative_stats['roas'].apply(get_creative_status)

    # NaN을 0으로 변환
    creative_stats = creative_stats.fillna(0)

    return creative_stats.to_dict('records')


def _calculate_metrics_inmemory(df):
    """
    In-Memory 방식으로 지표 계산 (데이터베이스 없이)

    Args:
        df: pandas DataFrame with columns: date, campaign_name, spend, clicks, conversions, revenue, impressions (optional)

    Returns:
        dict: 계산된 메트릭스
    """
    import numpy as np

    # 전체 지표 계산
    total_spend = df['spend'].sum()
    total_revenue = df['revenue'].sum()
    total_clicks = df['clicks'].sum()
    total_conversions = df['conversions'].sum()
    total_impressions = df['impressions'].sum() if 'impressions' in df.columns else 0

    # 계산 지표
    avg_roas = round(total_revenue / total_spend, 2) if total_spend > 0 else 0
    avg_ctr = round((total_clicks / total_impressions * 100), 2) if total_impressions > 0 else 0
    avg_cpc = round(total_spend / total_clicks, 0) if total_clicks > 0 else 0
    avg_cpa = round(total_spend / total_conversions, 0) if total_conversions > 0 else 0
    cvr = round((total_conversions / total_clicks * 100), 2) if total_clicks > 0 else 0
    avg_order_value = round(total_revenue / total_conversions, 0) if total_conversions > 0 else 0

    # 캠페인별 통계 (광고유형 포함)
    agg_dict = {
        'spend': 'sum',
        'revenue': 'sum',
        'clicks': 'sum',
        'conversions': 'sum',
        'impressions': 'sum' if 'impressions' in df.columns else lambda x: 0
    }

    # ad_type이 있으면 그룹화에 포함
    if 'ad_type' in df.columns:
        # 캠페인별 ad_type은 첫 번째 값 사용 (동일 캠페인은 동일 유형이라고 가정)
        campaign_stats = df.groupby('campaign_name').agg({
            **agg_dict,
            'ad_type': 'first'
        }).reset_index()
    else:
        campaign_stats = df.groupby('campaign_name').agg(agg_dict).reset_index()
        campaign_stats['ad_type'] = 'sales'  # 기본값

    # 공통 지표 계산
    campaign_stats['roas'] = (campaign_stats['revenue'] / campaign_stats['spend']).replace([np.inf, -np.inf], 0).round(2)
    campaign_stats['cpl'] = (campaign_stats['spend'] / campaign_stats['conversions']).replace([np.inf, -np.inf], 0).round(0)
    campaign_stats['cpa'] = campaign_stats['cpl']  # CPA = CPL (같은 계산)

    if 'impressions' in df.columns and campaign_stats['impressions'].sum() > 0:
        campaign_stats['ctr'] = (campaign_stats['clicks'] / campaign_stats['impressions'] * 100).round(2)
    else:
        campaign_stats['ctr'] = 0

    campaign_stats['cvr'] = (campaign_stats['conversions'] / campaign_stats['clicks'] * 100).replace([np.inf, -np.inf], 0).round(2)

    # 광고유형별 주요지표 및 정렬
    def get_primary_metric(row):
        if row['ad_type'] == 'lead':
            return 'CPL'
        else:
            return 'ROAS'

    def get_primary_value(row):
        if row['ad_type'] == 'lead':
            return row['cpl']
        else:
            return row['roas']

    campaign_stats['primary_metric'] = campaign_stats.apply(get_primary_metric, axis=1)
    campaign_stats['primary_value'] = campaign_stats.apply(get_primary_value, axis=1)

    # 광고유형별로 분리하여 정렬 후 합치기
    sales_campaigns = campaign_stats[campaign_stats['ad_type'] == 'sales'].sort_values('roas', ascending=False)
    lead_campaigns = campaign_stats[campaign_stats['ad_type'] == 'lead'].sort_values('cpl', ascending=True)

    # 매출형 먼저, 잠재고객 나중에 (각각 순위 부여)
    sales_campaigns['rank'] = range(1, len(sales_campaigns) + 1)
    lead_campaigns['rank'] = range(1, len(lead_campaigns) + 1)

    campaign_stats = pd.concat([sales_campaigns, lead_campaigns], ignore_index=True)

    # 상태 판정 제거 (순위로 대체)
    campaign_stats['status'] = 'normal'  # 상태 배지 사용 안 함

    # 일별 트렌드 - 날짜별 전체 합계로 집계 (차트 표시용)
    daily = df.groupby('date').agg({
        'spend': 'sum',
        'revenue': 'sum',
        'clicks': 'sum',
        'conversions': 'sum',
        'impressions': 'sum' if 'impressions' in df.columns else lambda x: 0
    }).reset_index()

    daily['roas'] = (daily['revenue'] / daily['spend']).replace([np.inf, -np.inf], 0).round(2)

    if 'impressions' in df.columns:
        daily['ctr'] = (daily['clicks'] / daily['impressions'] * 100).replace([np.inf, -np.inf], 0).round(2)
    else:
        daily['ctr'] = 0

    daily['cvr'] = (daily['conversions'] / daily['clicks'] * 100).replace([np.inf, -np.inf], 0).round(2)

    # NaN을 0으로 변환
    daily = daily.fillna(0)
    campaign_stats = campaign_stats.fillna(0)

    # 날짜를 문자열로 변환 (Chart.js 호환성)
    daily['date'] = daily['date'].astype(str)

    # campaign_name을 보존한 원본 데이터 생성 (캠페인 분석용)
    daily_data_columns = ['date', 'campaign_name', 'spend', 'revenue', 'clicks', 'conversions']
    if 'impressions' in df.columns:
        daily_data_columns.append('impressions')
    if 'ad_type' in df.columns:
        daily_data_columns.append('ad_type')

    daily_data = df[daily_data_columns].copy()
    daily_data['date'] = daily_data['date'].astype(str)
    daily_data = daily_data.fillna(0)

    # 소재별 분석 (ad_creative_name 컬럼이 있는 경우)
    creatives = []
    if 'ad_creative_name' in df.columns:
        creatives = _calculate_creative_metrics(df)

    metrics = {
        # 기본 지표
        'total_spend': float(total_spend),
        'total_revenue': float(total_revenue),
        'total_clicks': int(total_clicks),
        'total_conversions': int(total_conversions),
        'total_impressions': int(total_impressions),

        # 계산 지표
        'avg_roas': avg_roas,
        'avg_ctr': avg_ctr,
        'avg_cpc': avg_cpc,
        'avg_cpa': avg_cpa,
        'cvr': cvr,
        'avg_order_value': avg_order_value,

        # 캠페인별 통계
        'campaigns': campaign_stats.to_dict('records'),

        # 일별 트렌드 (차트용 - 날짜별 집계)
        'daily_trend': daily.to_dict('records'),

        # 일별 상세 데이터 (캠페인 분석용 - campaign_name 포함)
        'daily_data': daily_data.to_dict('records'),

        # 소재별 통계 (있는 경우만)
        'creatives': creatives,
    }

    return metrics
